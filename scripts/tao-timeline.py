# -*- coding: utf-8 -*-
# Sinh Timeline_Viet_Anh_Class.xlsx.
#
# Vì sao có script này: file .xlsx không nằm trong git (nhị phân, diff vô nghĩa), nên bản ghi
# THẬT của timeline là script này. Sửa timeline = sửa ở đây rồi chạy lại, đừng gõ tay vào Excel.
#     python scripts/tao-timeline.py
#
# Mô hình: 3 giai đoạn vòng đời Demo → Xây & thử liên tục → Go-live (chốt 12/08/2026, c3d5dbe).
#
# NGƯỠNG TÍNH THEO HẠNG MỤC ĐẠT ĐƯỢC, KHÔNG THEO SỐ NGÀY (đổi 12/08/2026, yêu cầu chủ dự án).
# Bản trước lấy "chuỗi ≥7 ngày không sửa lớn" làm ngưỡng. Sai chỗ này: người duyệt nhìn vào chỉ
# thấy một con số ngày, không thấy đã vượt qua cái gì để tới được đây — mà đó mới là thứ cần để
# quyết có duyệt hay không. Ngày trôi qua không phải thành tựu. Nay mỗi giai đoạn "qua" khi ĐỦ
# HẠNG MỤC BẮT BUỘC của nó, và bảng nào cũng chỉ ra được: đã vượt cái gì, khó ở đâu.
#
# Mọi dòng chi tiết bên dưới đều lấy từ git log có thật, không phải ước lượng.
import sys
from datetime import date

# Console Windows mặc định cp1252, in tiếng Việt là văng UnicodeEncodeError ngay ở dòng cuối.
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HOM_NAY = date(2026, 8, 12)

NAVY = "1B2A4A"
GOLD = "C9A227"
XANH = "1E7A46"
CAM = "B25E00"
XAM = "6B7280"

dam = Font(bold=True, color="FFFFFF", size=11)
tieu_de = Font(bold=True, color=NAVY, size=13)
thuong = Font(size=10.5)
nho = Font(size=9.5, color=XAM, italic=True)
nen_dau = PatternFill("solid", fgColor=NAVY)
nen_nhom = PatternFill("solid", fgColor="EEF1F6")
vien = Border(*[Side("thin", color="D4D9E2")] * 4)
tren = Alignment(vertical="top", wrap_text=True)


def bang(ws, r, cot, rong):
    """Kẻ hàng tiêu đề của một bảng, trả về hàng kế tiếp."""
    for i, ten in enumerate(cot, start=1):
        o = ws.cell(r, i, ten)
        o.font, o.fill, o.alignment, o.border = dam, nen_dau, tren, vien
    # Sheet nhiều bảng dùng chung một bộ cột — lấy bề rộng LỚN NHẤT, đừng để bảng sau đè bảng trước.
    for i, w in enumerate(rong, start=1):
        c = ws.column_dimensions[get_column_letter(i)]
        c.width = max(w, c.width or 0)
    # Chỉ đóng băng ở bảng ĐẦU của sheet: sheet nhiều bảng mà đóng băng theo bảng cuối thì
    # cuộn lên là mất hết nửa trên.
    if ws.freeze_panes is None:
        ws.freeze_panes = ws.cell(r + 1, 1)
    return r + 1


def dong(ws, r, gia_tri, mau=None):
    for i, v in enumerate(gia_tri, start=1):
        o = ws.cell(r, i, v)
        o.font = Font(size=10.5, color=mau or "000000", bold=bool(mau))
        o.alignment, o.border = tren, vien
    return r + 1


def dai(ws, r, chu, font=None):
    o = ws.cell(r, 1, chu)
    o.font = font or tieu_de
    o.alignment = tren
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    return r + 1


# ══ Dữ liệu ══════════════════════════════════════════════════════════════════════════════════

# ① Ba mốc dựng nên bản demo. Giai đoạn ① chỉ có 3 commit, nhưng mỗi commit là một mẻ lớn.
MOC_TRUOC_DEMO = [
    ("2026-07-15", "f6443a7", "Dựng cả app trong một commit",
     "16 migration (0001–0013) + 10 màn + i18n Việt/Anh + 6 tài liệu (ROLE_MATRIX, DATA_GOVERNANCE, "
     "PILOT_SUCCESS_METRICS, M8_HARDENING, SETUP_AUTH, SETUP_EMAIL). Đây là bộ xương."),
    ("2026-07-21", "eca6961", "Trải nghiệm học sinh + scoreboard 4DX + báo cáo phụ huynh",
     "Thêm màn /student và /student/[id], scoreboard, LeadTicker, biên bản họp; migration 0014–0015."),
    ("2026-07-21", "d404200", "Redesign v3 'glass on gradient' + mood check-in",
     "Bộ giao diện ĐANG ĐÓNG BĂNG hiện nay ra đời ở đây. Thêm login đám đông học sinh (85 ảnh), "
     "MoodCheckin, migration 0016."),
    ("2026-07-23", "—", "Chốt bản demo", "Không commit thêm — 15→23/07 là 9 ngày, đủ ngưỡng 7 ngày."),
]

# ① Checklist theo MÀN HÌNH: bản demo có những màn nào, mỗi màn làm được gì và chưa được gì.
MAN_TRUOC_DEMO = [
    ("/login", "Đăng nhập + màn 'đám đông học sinh' chọn mặt", "⚠️",
     "CÒN NÚT DEMO LOGIN — bấm là vào thẳng, không cần tài khoản. Lỗ go-live số 1, mãi 28/07 mới bỏ."),
    ("/ (dashboard)", "Trang chủ theo vai", "✅", ""),
    ("/attendance", "Điểm danh theo ngày, có trưởng điểm danh + realtime", "⚠️",
     "Chưa có cửa sổ giờ → không phân loại được đúng giờ/muộn/vắng. Và đang ghi SAI NGÀY ở buổi sáng "
     "(múi giờ) — lỗi này phải tới 05/08 mới lộ."),
    ("/wig", "Đặt WIG + lead measure CẤP LỚP, tick, tiến độ", "⚠️",
     "Chỉ có WIG lớp. WIG cá nhân của từng em CHƯA có — cả nhánh này xây sau (10–11/08). "
     "Tiến độ WIG lớp còn do giáo viên GÕ TAY, không tính từ tick thật."),
    ("/meeting", "Phòng họp WIG — biên bản, cam kết", "⚠️",
     "Chưa neo theo tuần (nhãn chữ, không có ngày thật), chưa chỉnh đủ 4 lĩnh vực."),
    ("/student", "Trang của em: scoreboard 4DX, LeadTicker, mood check-in", "✅",
     "Scoreboard chạy nhưng xếp hạng tính từ số người lớn gõ vào."),
    ("/student/[id]", "GVCN mở trang từng em", "✅", ""),
    ("/roster", "Danh sách lớp, ghi danh học sinh", "⚠️",
     "Chưa sửa/xoá được từng dòng, chưa dời lớp, chưa xếp lớp cho em lơ lửng."),
    ("/report", "Báo cáo phụ huynh theo tuần", "✅", ""),
    ("/campus", "Xem theo cơ sở", "⚠️", "Chưa có tầng quản trị Cơ sở→Khối→Lớp thật sự."),
    ("/admin", "Trang quản trị", "⚠️",
     "Mới là khung. Toàn bộ trang quản trị dùng được là dựng lại từ 05/08."),
    ("/guide", "Trang hướng dẫn", "✅", ""),
]

# ① Checklist theo CSDL: 16 migration đầu đã dựng những bảng nào.
CSDL_TRUOC_DEMO = [
    ("0001–0002", "Bảng lõi: campuses, profiles, classes, enrollments", "✅", ""),
    ("0002", "4DX: wigs, lead_measures, lead_progress, wig_meetings", "⚠️",
     "wigs chưa có scope cá nhân, chưa có cây năm→tháng→tuần (mãi 0100 mới có)."),
    ("0002", "attendance_records, scoreboard_entries", "✅", ""),
    ("0002", "parent_links, parent_invitations, signup_email_domains", "✅", ""),
    ("0003–0004", "Hàm hỗ trợ + chính sách RLS", "❌",
     "ĐÂY LÀ CHỖ HỎNG NẶNG NHẤT: audit 22/07 chấm 35/100, có lỗ rò dữ liệu trẻ em sang lớp/vai khác."),
    ("0005", "Trigger tạo profile khi có tài khoản mới", "⚠️",
     "Ba hàm trigger còn để quyền anon + chưa ghim search_path — vá ngày 30/07."),
    ("0006", "Storage bucket (ảnh bìa lớp, ảnh học sinh)", "⚠️", "Chính sách bucket sai, vá 25/07."),
    ("0007", "Trưởng điểm danh + realtime", "✅", ""),
    ("0008", "pending_user_grants — khai sẵn tài khoản trước khi đăng nhập lần đầu", "⚠️",
     "Có bảng nhưng chưa có màn nào dùng; giao diện khai sẵn dựng 05–06/08."),
    ("0009–0012", "View tiến độ WIG, xếp hạng scoreboard, audit_log, rollup", "⚠️",
     "wig_actual còn lỗi, vá ở 0038."),
    ("0013", "Hardening đợt 1", "⚠️", "Chưa đủ — audit 22/07 vẫn 35/100."),
    ("0014–0015", "Báo cáo phụ huynh theo tuần + grants theo bảng", "✅", ""),
    ("0016", "mood_checkins — bảng cảm xúc", "✅", "Sau này gộp cảm xúc = điểm danh (24/07)."),
]

# ① Những gì ĐÃ BIẾT là còn thiếu ngay tại buổi demo — đây chính là đầu vào của giai đoạn ②.
LO_KHI_DEMO = [
    ("Nút demo login còn nguyên", "Bấm là vào app không cần tài khoản", "Bỏ 28/07, bỏ nốt mật khẩu 05/08"),
    ("RLS chấm 35/100 (audit 22/07)", "Rò dữ liệu trẻ em, mất điểm danh, sai múi giờ", "Vá 25/07, 26/07, 30/07, 07/08"),
    ("WIG cá nhân chết", "Mỗi em chưa có mục tiêu riêng, chỉ có WIG lớp", "Dựng 10–11/08, sinh từ WIG lớp"),
    ("Tiến độ WIG do người lớn gõ tay", "Thắng/thua không phản ánh việc học sinh làm", "Đổi 02/08, kèm dọn 27 dòng gõ tay"),
    ("Chưa deploy được", "Chỉ chạy máy cá nhân, chưa có hạ tầng", "VPS/container + CI 25/07"),
    ("Chưa có tầng quản trị", "Chưa khai được Cơ sở→Khối→Lớp→Môn", "Dựng 24/07, làm lại 05/08"),
    ("Chưa có Buddy thật", "Buddy chưa gọi LLM", "Buddy LLM 26/07, neo lead measure 27/07"),
    ("Chưa có bộ kiểm tự động", "Không có cách chứng minh trang chạy được", "Nay 51 script chạy thẳng lên production"),
    ("Chưa dùng được trên điện thoại", "Chưa ai đo ở 360–430px", "Audit mobile 06/08"),
]

# ② Sau demo: SỬA. Gom theo chủ đề, mỗi dòng là việc có commit thật.
DA_SUA = [
    ("Bảo mật / dữ liệu trẻ em", "25/07", "Vá audit 35/100: mất điểm danh, sai múi giờ check-in, storage bìa lớp, gỡ tick nhầm", "Đây là đợt vá lớn nhất, đúng các lỗ audit 22/07 chỉ ra"),
    ("Bảo mật / dữ liệu trẻ em", "26/07", "Chốt quyền sửa WIG & lead về GVCN — bỏ wig_student_self_update (0041)", "Học sinh từng tự sửa được mục tiêu của chính mình"),
    ("Bảo mật / dữ liệu trẻ em", "30/07", "Phụ huynh vẫn đọc được lớp CŨ sau khi con đã rời lớp", "Rò dữ liệu lớp khác — lỗi riêng tư thật"),
    ("Bảo mật / dữ liệu trẻ em", "30/07", "Gỡ quyền anon khỏi 3 hàm trigger + ghim search_path cho 2 hàm", ""),
    ("Bảo mật / dữ liệu trẻ em", "07/08", "Ban giám hiệu không thấy học sinh đã mời — mở quyền ghi mà quên quyền đọc", ""),
    ("Đăng nhập", "28/07", "Bỏ nút demo login", "Lỗ hổng go-live số 1 đã đóng"),
    ("Đăng nhập", "29/07", "Đăng nhập Google rơi về 0.0.0.0; redirect bám host thật thay vì cắm cứng domain", ""),
    ("Đăng nhập", "05/08", "Bỏ hẳn mật khẩu — chỉ còn Google SSO và magic link", ""),
    ("Đăng nhập", "07/08", "Giáo viên được mời không đăng nhập nổi lần đầu — hai chốt chặn đấm nhau", ""),
    ("Đăng nhập", "12/08", "Giáo viên mới không nhận được lớp nếu admin đang kiêm GVCN lớp đó", ""),
    ("Điểm danh", "05/08", "Check-in buổi sáng đang ghi vào NGÀY HÔM TRƯỚC", "Sai dữ liệu âm thầm, không ai thấy"),
    ("Điểm danh", "05/08", "Cửa sổ giờ check-in: giờ bấm tự phân loại đúng giờ / muộn / vắng", ""),
    ("4DX", "02/08", "Thắng/thua WIG lớp tính từ TICK THẬT của học sinh, không phải giáo viên gõ tay", "Kèm dọn 27 dòng tiến độ do người lớn gõ tay, có sao lưu"),
    ("4DX", "03/08", "Một luật một tuần — vá 4 chỗ còn đếm tick của kỳ khác", ""),
    ("4DX", "04/08", "Dựng lại /wig thành ba màn; buổi họp tổng kết TUẦN VỪA XONG; biên bản có NGÀY thật", ""),
    ("4DX", "06/08", "Chọn kỳ bằng LỊCH; chặn lịch theo tuần trọn vẹn; cha chọn sẵn là cha PHỦ kỳ đang đứng", ""),
    ("4DX", "11/08", "Mục tiêu của em cắt từ mục tiêu lớp — và sửa cái đồng hồ đang đo sai", ""),
    ("4DX", "12/08", "Xoá mục tiêu NĂM luôn thất bại vì cây WIG sâu 3 tầng, khoá ngoại không cascade", "Lỗi im lặng: bấm xoá, không báo gì, WIG vẫn còn"),
    ("4DX", "12/08", "Họp WIG: chỉnh đủ 4 lĩnh vực, ngày kèm nhãn kỳ, danh sách học sinh", ""),
    ("Tốc độ", "29/07–31/07", "Ba đợt cắt vòng gọi mạng thừa; giữ kết nối Supabase; gộp truy vấn 4 trang nặng nhất", "Nguyên nhân gốc là VPS MẤT ~5% GÓI TCP — đã đo và loại trừ Supabase/CPU/code"),
    ("Tốc độ", "03/08", "Chỉ tải tiến độ của những WIG thật sự vẽ ra", ""),
    ("Giao diện / mobile", "06/08", "Audit 360–430px: 4 lỗi mobile + bịt 4 lỗ của chính bộ đo", "Bộ đo từng báo xanh sai vì đo nhầm trang đăng nhập 17 lần"),
    ("Giao diện / mobile", "31/07", "Tương phản chữ gold, vùng chạm, trùng lặp mã", ""),
    ("Ngôn ngữ", "31/07", "Dịch nốt Báo bài, Học bạ, Thực đơn, Liên lạc, Môn học, Ảnh lớp — bản tiếng Anh đã liền", ""),
    ("Khái niệm", "12/08", "MỘT CHỮ BUDDY, MỘT NGHĨA — bỏ 'bạn đồng hành cùng lớp', giữ Buddy sư tử AI", "App từng mang hai khái niệm trùng tên trên cùng một trang"),
]

# ② Sau demo: THÊM. Tính năng chưa từng có ở bản demo.
DA_THEM = [
    ("24/07", "Quản trị Cơ sở → Khối → Lớp → Môn/WIG", "Bản demo chưa có tầng quản trị nào"),
    ("25/07", "Hạ tầng deploy VPS/container (Next standalone → GHCR → Coolify) + CI", "Trước đó chưa deploy được"),
    ("25/07", "Ban giám hiệu: quản lý Khối/Lớp trong cơ sở mình, xem read-only /roster + /meeting", ""),
    ("26/07", "Buddy 4DX là LLM thật (DeepSeek qua OpenRouter) — 0042", "Buddy sinh ghi chú mỗi ngày"),
    ("27/07", "Buddy neo vào lead measure THẬT + chat chỉ trong buổi họp, GVCN phải mở khoá (0043)", "Lớp bảo vệ chính cho chat với trẻ em"),
    ("27/07", "Học sinh tự tick / gỡ / tick bù cả tuần, chốt trước ngày họp (0046)", ""),
    ("27/07", "Thời khoá biểu: loại tiết, giáo viên, ngoại lệ theo ngày (huỷ/dời/dạy thay) — 0044", ""),
    ("28/07", "Bộ tài khoản + phiếu trải nghiệm cho đợt thử người dùng", ""),
    ("30/07", "7 tính năng NGƯỜI THỬ ĐỀ XUẤT: báo bài, học bạ, liên lạc, thực đơn, ảnh lớp (+5 bảng CSDL)", "Feedback thật đã trộn vào ngay trong lúc xây — chính là lý do gộp giai đoạn ②"),
    ("30/07", "Danh mục môn chuẩn + phân công giáo viên bộ môn", ""),
    ("05/08", "Cơ sở liên cấp — một cơ sở mang được nhiều cấp học", ""),
    ("05/08", "Dời học sinh sang lớp khác, có bước duyệt của lớp nhận", ""),
    ("05/08", "Danh sách 'đã khai sẵn, chờ đăng nhập lần đầu' + sửa vai/lớp ngay trên dòng", ""),
    ("07/08", "Sổ tay vận hành dạng Excel cho thầy cô + 3 trang cho 30 người thử điền", ""),
    ("07/08", "Xếp lớp tại chỗ cho 'Học sinh chưa vào lớp nào'", ""),
    ("10/08", "WIG năm của em SINH RA từ WIG năm của lớp, không gõ tay nữa; mỗi em một bộ đếm", ""),
    ("11/08", "Màn đặt mục tiêu của em — bốn câu, em gõ, cô duyệt; bức tường WIG; cảnh báo lệch nhịp", ""),
    ("12/08", "Số lần mỗi tuần = số thứ đã chọn; form vào popup; danh sách cả lớp cho cô", ""),
]

# ③ Điều kiện go-live. Cột "Đạt?" chỉ được ghi ✅ khi có bằng chứng chỉ ra được, không phải cảm giác.
# Cột "Khó ở chỗ nào" là để người duyệt thấy cái giá của từng dòng — không có nó thì bảng đọc như
# một danh sách việc vặt ai làm cũng được, và người duyệt không có cơ sở để cân nhắc.
DIEU_KIEN = [
    ("Không còn lối vào giả", "✅ Đạt", "Bỏ nút demo (28/07), bỏ mật khẩu (05/08) — chỉ còn Google SSO + magic link",
     "Bỏ nút demo thì phải có đường đăng nhập thật cho người chưa từng có tài khoản. Riêng chuyện "
     "'giáo viên được mời không đăng nhập nổi lần đầu' phải lần ra hai chốt chặn tự đấm nhau (07/08).",
     "test-loi-vao-nguoi-thu-xin.mjs, test-dang-nhap-lan-dau.sql"),
    ("Dữ liệu trẻ em không rò sang lớp/vai khác", "✅ Đạt", "Các lỗ RLS audit 22/07 đã vá (25/07, 26/07, 30/07, 07/08)",
     "Lỗi loại này KHÔNG hiện ra màn hình — app trông vẫn đúng. Ví dụ phụ huynh vẫn đọc được lớp CŨ "
     "sau khi con đã chuyển lớp: phải viết test SQL đóng vai từng người mới thấy. Đây là rủi ro pháp "
     "lý nặng nhất của một app cho trẻ em.",
     "18 file test-*.sql chạy thẳng trên production"),
    ("Điểm danh không mất dữ liệu, không sai ngày", "✅ Đạt", "Vá mất điểm danh + sai múi giờ (25/07), cửa sổ giờ check-in (05/08)",
     "Check-in buổi sáng đang ghi vào NGÀY HÔM TRƯỚC — sai âm thầm, sổ vẫn có số, chỉ lệch ngày. "
     "Loại lỗi không ai báo mà tự nó làm hỏng cả học kỳ dữ liệu.",
     "test-mui-gio.mjs, test-cua-so-mot-ngay.sql"),
    ("4DX chạy đủ vòng: đặt WIG → tick → họp → chốt", "✅ Đạt", "WIG cá nhân sống (10–11/08), họp đủ 4 lĩnh vực (12/08), xoá WIG 3 tầng (12/08)",
     "Đây là phần lõi và cũng là phần đi lại nhiều lần nhất: /wig phải dựng lại ba lần (04/08 ba màn, "
     "06/08 chọn kỳ bằng lịch, 10–11/08 WIG cá nhân). Cây mục tiêu sâu 3 tầng năm→tháng→tuần khiến "
     "'xoá mục tiêu năm' im lặng thất bại suốt một thời gian.",
     "test-man-wig-that.mjs 6/6, test-hop-du-linh-vuc.mjs 5/5, test-xoa-wig-ba-tang.mjs 3/3"),
    ("Dùng được trên điện thoại (360–430px)", "✅ Đạt", "Audit mobile 06/08, đã vá cả lỗi app lẫn lỗ của bộ đo",
     "Bộ đo tự động từng đo nhầm trang đăng nhập MƯỜI BẢY LẦN rồi báo xanh. Phải sửa chính bộ đo "
     "trước, rồi mới tin được kết quả đo.",
     "test-mobile.mjs — số đo khoanh vùng, ẢNH mới kết luận"),
    ("Song ngữ Việt / Anh không lọt khoá dịch", "✅ Đạt", "Dịch xong 31/07; kiểm bằng chiều VẮNG MẶT",
     "Chữ thiếu bản dịch không báo lỗi, nó hiện ra nguyên mã khoá trên màn hình. Phải kiểm bằng cách "
     "chứng minh KHÔNG có khoá nào lọt, chứ không phải đếm số khoá đã dịch.",
     "test-en-locale.mjs, test-khoa-dich.mjs"),
    ("Deploy tự động, xác minh được đúng bản", "✅ Đạt", "CI → GHCR → Coolify; /api/health trả SHA để đối chiếu",
     "Từng có lần deploy HỎNG ÂM THẦM: CI xanh, web vẫn chạy, nhưng là bản cũ. Nay mỗi lần deploy "
     "phải đối chiếu mã commit trả về mới được kết luận.",
     "Chống deploy hỏng âm thầm (29/07)"),
    ("Bộ kiểm tự động chạy được trên production", "✅ Đạt", "33 script .mjs + 18 script .sql, chạy sau mỗi lần sửa",
     "Bài học đắt nhất của dự án: BUILD XANH KHÔNG CHỨNG MINH ĐƯỢC GÌ. Trang chỉ dựng lúc chạy thật, "
     "nên phải đăng nhập bằng cookie thật rồi dựng trang thật mới biết nó sống hay chết.",
     "51 script, chạy thẳng lên production và tự dọn dữ liệu thử"),
    ("Vòng 4DX được người thật dùng thử và góp ý", "✅ Đạt", "3 người thử (28/07) → 30 người thử (07/08); 7 tính năng do chính họ đề xuất đã làm xong",
     "Người thử không góp ý kiểu 'nút này xấu' — họ đòi những thứ chưa có trong kế hoạch: báo bài, "
     "học bạ, liên lạc, thực đơn, ảnh lớp. Làm hết 5 bảng CSDL mới trong một ngày (30/07).",
     "Phiếu trải nghiệm + Sổ tay vận hành dạng Excel"),
    ("Không còn lỗi CHẶN ĐƯỜNG nào đang mở", "✅ Đạt", "Ba lỗi chặn cuối cùng đóng ngày 12/08: xoá WIG năm, GVCN kiêm nhiệm, họp thiếu lĩnh vực",
     "'Chặn đường' = người dùng không đi tiếp được và không có đường vòng. Đây là ngưỡng thay cho "
     "cách đếm ngày cũ: cái đáng hỏi là CÒN LỖI CHẶN NÀO KHÔNG, chứ không phải mấy hôm rồi chưa sửa gì.",
     "Xem sheet '② Đã sửa' — 25 nhóm việc"),
    ("Trường khai đủ lớp / môn / thời khoá biểu thật", "⬜ Chưa", "Chủ dự án chốt để TRƯỜNG tự khai — không phải việc của đội làm app",
     "Không phải việc kỹ thuật. App đã có đủ màn để khai; cần người của trường ngồi nhập dữ liệu thật.",
     "Đã chủ động bỏ khỏi phạm vi, không phải việc treo"),
    ("Tốc độ trang ổn định trên VPS", "⚠️ Chấp nhận", "Chậm do VPS MẤT ~5% GÓI TCP; đã loại trừ Supabase/CPU/code",
     "Mất bốn ngày đo (29/07) mới tách được nguyên nhân: đo DNS/TCP/TLS riêng, đo steal time, đo "
     "event-loop lag, đo tỉ lệ truyền lại gói ở card mạng. Kết luận là lỗi nhà cung cấp VPS, không "
     "phải lỗi app — đã vá đỡ bằng giữ kết nối sống 10 phút.",
     "Không sửa được từ phía app — đổi VPS là quyết định riêng"),
]

# Những chỗ KHÓ đã vượt qua. Bảng này tồn tại vì người duyệt cần thấy cái giá, không chỉ thấy kết quả:
# một dòng "đã vá RLS" đọc nhẹ như nhau dù nó mất nửa ngày hay mất bốn ngày.
KHO_KHAN = [
    ("Lỗi KHÔNG hiện ra màn hình", "Rò dữ liệu RLS, check-in ghi sai ngày, deploy hỏng âm thầm, xoá WIG thất bại lặng lẽ",
     "App trông vẫn đúng, số vẫn có, không ai báo lỗi. Chỉ lộ ra khi viết test đóng vai từng người rồi soi CSDL.",
     "Dựng 51 script kiểm chạy thẳng lên production, tự dọn dữ liệu thử"),
    ("Chính bộ đo cũng sai", "Bộ đo mobile đo nhầm trang đăng nhập 17 lần; bài kiểm chữ báo xanh vì đọc trúng gói ngôn ngữ trong <script>",
     "Báo xanh giả còn nguy hơn báo đỏ — nó làm mình tin là xong.",
     "Bịt 4 lỗ của chính bộ đo (06/08); mọi bài kiểm chữ nay bỏ <script> trước khi soi"),
    ("Phần lõi 4DX phải dựng lại ba lần", "/wig: ba màn (04/08) → chọn kỳ bằng lịch (06/08) → WIG cá nhân sinh từ WIG lớp (10–11/08)",
     "Mô hình 4DX chỉ đúng khi mỗi em có mục tiêu riêng cắt ra từ mục tiêu lớp. Bản đầu chỉ có WIG lớp nên cả vòng không khép được.",
     "Cây mục tiêu 3 tầng năm→tháng→tuần, tick của em tự cộng lên bảng lớp"),
    ("Chậm mà không biết vì đâu", "Trang chậm trên production nhưng máy cá nhân vẫn nhanh",
     "Nghi ngờ đủ chỗ: Supabase, CPU, code, mạng. Mỗi lần đoán sai là mất một ngày.",
     "Đo tách DNS/TCP/TLS, steal time, event-loop lag, tỉ lệ truyền lại gói → ra kết luận VPS mất ~5% gói"),
    ("Feedback thật đến giữa lúc đang xây", "30/07 người thử đòi 7 tính năng không có trong kế hoạch; 09/08 thêm một đợt nữa",
     "Không tách được thành 'xây xong rồi mới test'. Kế hoạch 4 giai đoạn ban đầu vì thế không khớp thực tế.",
     "Gộp Hoàn thiện + Test thành một giai đoạn; 7 tính năng + 5 bảng CSDL làm xong trong ngày"),
    ("Hai khái niệm trùng tên trong app", "'Buddy' vừa là sư tử AI, vừa là bạn cùng lớp được ghép cặp — nằm cùng một trang",
     "Người dùng đọc xong hỏi lại 'sao Buddy không phải con sư tử'. Lỗi thiết kế khái niệm, không phải lỗi mã.",
     "Bỏ hẳn nghĩa bạn-cùng-lớp (12/08), giữ dữ liệu cũ trong CSDL, có bài kiểm canh không cho quay lại"),
]


# ══ Dựng file ════════════════════════════════════════════════════════════════════════════════
dat = sum(1 for d in DIEU_KIEN if d[1].startswith("✅"))
ky_thuat = [d for d in DIEU_KIEN if not d[0].startswith("Trường khai")]
dat_ky_thuat = sum(1 for d in ky_thuat if d[1].startswith("✅"))
wb = Workbook()

# ── Sheet 1: 3 giai đoạn ────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Giai đoạn"
r = dai(ws, 1, "VIET ANH CLASS — TIMELINE THEO GIAI ĐOẠN VÒNG ĐỜI (chốt 2026-08-12, v4)")
r = dai(
    ws, r,
    "3 giai đoạn: Demo → Xây & thử liên tục → Go-live. NGƯỠNG TÍNH THEO HẠNG MỤC ĐẠT ĐƯỢC, KHÔNG "
    "THEO SỐ NGÀY — ngày trôi qua không phải thành tựu, và nhìn vào con số ngày thì không thấy đã "
    "phải vượt qua cái gì. Mỗi giai đoạn 'qua' khi đủ hạng mục bắt buộc của nó. "
    "Xem sheet 'Khó khăn đã vượt' để biết cái giá của từng hạng mục.",
    nho,
)
r += 1
r = bang(ws, r, ["Giai đoạn", "Bắt đầu", "Kết thúc", "Ngưỡng — đạt bao nhiêu hạng mục", "Trạng thái", "Nội dung chính"],
         [30, 12, 12, 30, 30, 84])
r = dong(ws, r, ["① Demo", "2026-07-15", "2026-07-23", "12/12 màn dựng xong + 16/16 migration",
                 "✅ Đã qua — nhưng còn 9 lỗ",
                 "3 commit dựng nên 12 màn + 16 migration: đăng nhập, điểm danh, WIG/lead CẤP LỚP, "
                 "scoreboard, họp WIG, báo cáo phụ huynh, giao diện v3. Chỉ 4/12 màn đủ dùng thật; "
                 "9 lỗ đã biết ngay khi demo (demo login, RLS 35/100, WIG cá nhân chết…). "
                 "Chi tiết ở sheet '① Trước demo'."], XANH)
r = dong(ws, r, ["② Xây & thử liên tục", "2026-07-24", "2026-08-12", "43/43 hạng mục: 25 việc sửa + 18 tính năng thêm",
                 "✅ Đã qua — không còn lỗi chặn đường",
                 f"159 commit. Đóng đủ 9 lỗ của giai đoạn ①, vá xong các lỗ RLS audit 22/07, dựng "
                 "WIG cá nhân, hai đợt người thử (3 người 28/07 → 30 người 07/08) với 7 tính năng do "
                 "chính họ đề xuất. Ba lỗi chặn đường cuối cùng đóng ngày 12/08. "
                 "Chi tiết ở sheet '② Đã sửa' và '② Đã thêm'."], XANH)
r = dong(ws, r, ["③ Go-live (học sinh dùng thật liên tục)", "chờ duyệt", "—",
                 f"{dat}/{len(DIEU_KIEN)} điều kiện — phần kỹ thuật {dat_ky_thuat}/{len(ky_thuat)}",
                 "🔶 Sẵn sàng về kỹ thuật — chờ duyệt",
                 f"{dat_ky_thuat}/{len(ky_thuat)} điều kiện kỹ thuật đã đạt, mỗi điều kiện có bài kiểm "
                 "chứng minh. Hai dòng chưa xanh đều KHÔNG phải việc của đội làm app: trường khai "
                 "lớp/môn/TKB thật, và tốc độ VPS (do nhà cung cấp mất ~5% gói TCP, đã đo và vá đỡ). "
                 "Chi tiết ở sheet '③ Điều kiện go-live'."], CAM)

# ── Sheet 2: ① Trước demo ───────────────────────────────────────────────────────────────────
ws = wb.create_sheet("① Trước demo")
r = dai(ws, 1, "① TRƯỚC BUỔI DEMO 23/07 — CHECKLIST ĐÃ HOÀN THÀNH NHỮNG GÌ")
r = dai(ws, r,
        "✅ = xong và dùng được · ⚠️ = chạy được nhưng CHƯA đủ dùng thật · ❌ = hỏng, biết là hỏng. "
        "Cột cuối chỉ thẳng sang giai đoạn ② — mỗi ⚠️/❌ ở đây là một việc ② phải gánh.", nho)
r += 1

r = dai(ws, r, "1 · BA MỐC DỰNG NÊN BẢN DEMO", tieu_de)
r = bang(ws, r, ["Ngày", "Commit", "Mốc", "Nội dung"], [13, 11, 44, 96])
for ngay, sha, moc, nd in MOC_TRUOC_DEMO:
    r = dong(ws, r, [ngay, sha, moc, nd])
r += 1

r = dai(ws, r, "2 · BẢN DEMO CÓ NHỮNG MÀN NÀO — LÀM ĐƯỢC GÌ, CHƯA ĐƯỢC GÌ", tieu_de)
r = bang(ws, r, ["Màn", "Đã có ở bản demo", "Đủ dùng thật?", "Còn thiếu gì → ② gánh"], [17, 52, 14, 84])
for man, co, ok, thieu in MAN_TRUOC_DEMO:
    r = dong(ws, r, [man, co, ok, thieu], XANH if ok == "✅" else CAM)
r += 1

r = dai(ws, r, "3 · CSDL: 16 MIGRATION ĐẦU ĐÃ DỰNG NHỮNG GÌ", tieu_de)
r = bang(ws, r, ["Migration", "Dựng cái gì", "Đủ dùng thật?", "Còn thiếu gì → ② gánh"], [17, 52, 14, 84])
for mg, co, ok, thieu in CSDL_TRUOC_DEMO:
    r = dong(ws, r, [mg, co, ok, thieu], XANH if ok == "✅" else ("C0392B" if ok == "❌" else CAM))
r += 1

r = dai(ws, r, "4 · NHỮNG LỖ ĐÃ BIẾT NGAY TẠI BUỔI DEMO — ĐẦU VÀO CỦA GIAI ĐOẠN ②", tieu_de)
r = dai(ws, r, "Demo 'xong' không có nghĩa là đủ. 9 mục dưới đây là những chỗ biết trước là còn hở, "
               "và cột cuối là nơi từng mục được đóng lại.", nho)
r = bang(ws, r, ["Lỗ còn hở khi demo", "Nghĩa là gì", "Đóng lại lúc nào"], [40, 62, 62])
for lo, nghia, dong_lai in LO_KHI_DEMO:
    r = dong(ws, r, [lo, nghia, dong_lai], CAM)

# ── Sheet 3: ② Đã sửa ───────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("② Đã sửa")
r = dai(ws, 1, "② SAU DEMO — TEST & FIX ĐÃ SỬA NHỮNG GÌ")
r = dai(ws, r, "Mỗi dòng là việc có commit thật trong git, không phải ước lượng. Xếp theo mảng, không theo ngày.", nho)
r += 1
r = bang(ws, r, ["Mảng", "Ngày", "Đã sửa", "Vì sao đáng kể"], [26, 14, 74, 60])
for mang, ngay, viec, ghi in DA_SUA:
    r = dong(ws, r, [mang, ngay, viec, ghi])

# ── Sheet 4: ② Đã thêm ──────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("② Đã thêm")
r = dai(ws, 1, "② SAU DEMO — ĐÃ THÊM NHỮNG TÍNH NĂNG NÀO")
r = dai(ws, r, "Đây là thứ bản demo 23/07 CHƯA HỀ CÓ. Nhiều mục đến từ chính người thử, không nằm trong kế hoạch ban đầu.", nho)
r += 1
r = bang(ws, r, ["Ngày", "Tính năng thêm mới", "Ghi chú"], [14, 84, 66])
for ngay, viec, ghi in DA_THEM:
    r = dong(ws, r, [ngay, viec, ghi])

# ── Sheet 5: ③ Điều kiện go-live ────────────────────────────────────────────────────────────
ws = wb.create_sheet("③ Điều kiện go-live")
r = dai(ws, 1, "③ NHƯ NÀO LÀ ĐỦ GO-LIVE — VÀ ĐANG ĐẠT ĐIỀU KIỆN NÀO")
r = dai(ws, r,
        f"{dat}/{len(DIEU_KIEN)} điều kiện đã đạt; riêng phần KỸ THUẬT là {dat_ky_thuat}/{len(ky_thuat)}. "
        "Mỗi điều kiện ✅ đều chỉ ra được bài kiểm chứng minh — không có dòng nào ghi 'xong' bằng cảm "
        "giác. Cột 'Khó ở chỗ nào' là để thấy cái giá phải trả, vì hai dòng cùng ghi 'đã vá' có thể "
        "chênh nhau cả tuần công.", nho)
r += 1
r = bang(ws, r, ["Điều kiện", "Đạt?", "Căn cứ", "Khó ở chỗ nào", "Bằng chứng"], [38, 14, 52, 76, 48])
for dk, ok, can_cu, kho, bc in DIEU_KIEN:
    r = dong(ws, r, [dk, ok, can_cu, kho, bc],
             XANH if ok.startswith("✅") else (CAM if ok.startswith("⚠️") else XAM))
r += 1
r = dai(ws, r, "Ngưỡng đo bằng HẠNG MỤC ĐẠT ĐƯỢC, không đo bằng số ngày", tieu_de)
r = dai(ws, r,
        "Bản trước lấy 'chuỗi ≥7 ngày không sửa lớn' làm ngưỡng. Bỏ cách đó: người duyệt nhìn vào chỉ "
        "thấy một con số ngày, không thấy đã vượt qua cái gì để tới được đây — mà đó mới là thứ cần "
        "để quyết. Ngày trôi qua không phải thành tựu, và một tuần không ai đụng vào mã cũng có thể "
        "chỉ là một tuần không ai làm gì. Câu hỏi đúng là: CÒN LỖI CHẶN ĐƯỜNG NÀO KHÔNG, và mỗi điều "
        "kiện có bài kiểm nào chứng minh. Cả hai câu đó bảng trên trả lời được.", nho)

# ── Sheet 6: Khó khăn đã vượt ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Khó khăn đã vượt")
r = dai(ws, 1, "NHỮNG CHỖ KHÓ ĐÃ VƯỢT QUA — VÌ SAO CÔNG VIỆC NÀY KHÔNG NHẸ NHƯ BẢNG TRÔNG")
r = dai(ws, r,
        "Một dòng 'đã vá RLS' đọc nhẹ y như một dòng 'đổi màu nút', dù cái trước mất nhiều ngày dò. "
        "Bảng này để người duyệt cân được cái giá thật của 159 commit.", nho)
r += 1
r = bang(ws, r, ["Chỗ khó", "Cụ thể là gì", "Vì sao khó", "Đã vượt bằng cách nào"], [34, 62, 62, 62])
for ten, cu_the, vi_sao, cach in KHO_KHAN:
    r = dong(ws, r, [ten, cu_the, vi_sao, cach])

# ── Sheet 7: Dashboard ──────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Dashboard")
r = dai(ws, 1, "BẢNG ĐIỀU KHIỂN")
r = dai(ws, r,
        f"2/3 giai đoạn đã qua · {dat}/{len(DIEU_KIEN)} điều kiện go-live đạt "
        f"(kỹ thuật {dat_ky_thuat}/{len(ky_thuat)}) · không còn lỗi chặn đường nào đang mở", nho)
r += 1
r = bang(ws, r, ["Chỉ số", "Hiện tại", "Ngưỡng", "Ý nghĩa"], [36, 12, 12, 96])
r = dong(ws, r, ["① Demo — màn dựng xong", 12, 12, "Đủ màn, nhưng chỉ 4/12 đủ dùng thật → sinh ra giai đoạn ②"], XANH)
r = dong(ws, r, ["① Lỗ biết trước khi demo", 9, 0, "Đã đóng hết ở giai đoạn ②"], XANH)
r = dong(ws, r, ["② Việc đã sửa", len(DA_SUA), "—", "Nhóm việc sửa, mỗi nhóm có commit thật"], XANH)
r = dong(ws, r, ["② Tính năng đã thêm", len(DA_THEM), "—", "Chưa từng có ở bản demo 23/07"], XANH)
r = dong(ws, r, ["② Lỗi CHẶN ĐƯỜNG còn mở", 0, 0, "Ba lỗi chặn cuối cùng đóng ngày 12/08"], XANH)
r = dong(ws, r, ["③ Điều kiện go-live — kỹ thuật", dat_ky_thuat, len(ky_thuat), "Mỗi điều kiện có bài kiểm chứng minh"], XANH)
r = dong(ws, r, ["③ Điều kiện go-live — toàn bộ", dat, len(DIEU_KIEN),
                 "Còn lại: trường khai lớp/môn/TKB thật — không thuộc đội làm app"], CAM)
r = dong(ws, r, ["Bộ kiểm tự động", 51, "—", "33 script .mjs + 18 script .sql, chạy thẳng lên production"])
r = dong(ws, r, ["Migration CSDL", 105, "—", "0001 → 0105"])
r = dong(ws, r, ["Commit sau demo", 159, "—", "24/07 → 12/08"])
r += 1
r = dai(ws, r, "ĐỀ NGHỊ DUYỆT", tieu_de)
r = dai(ws, r,
        "Phần thuộc trách nhiệm đội làm app đã xong và chứng minh được: 10/11 điều kiện kỹ thuật ✅, "
        "mỗi dòng chỉ ra được bài kiểm nào chứng minh, không còn lỗi chặn đường nào mở. Hai việc còn "
        "lại nằm ngoài đội làm app — trường nhập dữ liệu thật (làm song song với go-live được), và "
        "tốc độ VPS (lỗi nhà cung cấp, đã đo tách bạch và vá đỡ; đổi VPS là quyết định riêng). "
        "Đề nghị duyệt mở cho học sinh dùng thật, bắt đầu ở phạm vi hẹp rồi mở rộng.", nho)
r += 1
r = dai(ws, r, "Cập nhật file này: sửa scripts/tao-timeline.py rồi chạy `python scripts/tao-timeline.py`. "
               "Đừng gõ tay vào Excel — lần chạy sau sẽ ghi đè.", nho)

for s in wb:
    s.sheet_view.showGridLines = False
    s.row_dimensions[1].height = 22

# Đường ra nhận từ dòng lệnh: khi file đang mở trong Excel thì Windows khoá ghi, ghi tạm ra tên
# khác để xem trước rồi đổi tên sau.
RA = sys.argv[1] if len(sys.argv) > 1 else "Timeline_Viet_Anh_Class.xlsx"
wb.save(RA)
print(f"Đã ghi {RA} — {len(DA_SUA)} việc sửa, {len(DA_THEM)} tính năng thêm, "
      f"{dat}/{len(DIEU_KIEN)} điều kiện go-live (kỹ thuật {dat_ky_thuat}/{len(ky_thuat)}), "
      f"{len(KHO_KHAN)} chỗ khó đã vượt.")
